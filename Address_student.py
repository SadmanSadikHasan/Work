# Python3 code to select
# data from excel
import xlwings as xw
import openpyxl as xl
from openpyxl import load_workbook
import pandas as pd

# Specifying a sheet
ws = xw.Book("F:\Work\TABLE TEMPLATE\ADMISSION_STUDENTS - Copy.xlsx").sheets['Result 1']
XL = xl.load_workbook("F:\Work\TABLE TEMPLATE\ADMISSION_STUDENTS.xlsx")
sheet1 = XL["Result 1"]

# Selecting data from
# a single cell
v1 = ws.range("W3:W299").value
#v2 = ws.range("F5").value
#v1 = v1.replace('.', "")

v2 = len(v1)

for i in range(0, len(v1)):
    if v1[i] is not None:
        v1[i] = v1[i].replace(".", "")
print("Result:", v1)

for i in range(0, v2):
    sheet1.cell(row=i+3, column=23).value = v1[i]

XL.save('F:\Work\TABLE TEMPLATE\ADMISSION_STUDENTS.xlsx')
