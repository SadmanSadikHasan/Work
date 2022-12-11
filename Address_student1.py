import openpyxl

xl = openpyxl.load_workbook('F:\Work\DATASET\Details Data Final Fall-2007.xls')
sheet1 = xl['CSE']

mr = sheet1.max_row

rangeselected = []
for i in range(2,mr,1):
    rangeselected.append(sheet1.cell(row=i,column = 2).value)
for i in range(2,mr):
    print(rangeselected[i-1])