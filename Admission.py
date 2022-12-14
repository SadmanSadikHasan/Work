import xlwings as xw
import openpyxl as xl
from openpyxl import load_workbook
import pandas as pd

ws = xw.Book("F:\Work\DATASET\Details Data Final Fall-2007.xls").sheets['CSE']
#ws1 = xl.Book("F:\Work\TABLE TEMPLATE\ADMISSION_STUDENTS.xlsx")
XL = xl.load_workbook("F:\Work\TABLE TEMPLATE\ADMISSION_STUDENTS.xlsx")
sheet1 = XL["Result 1"]

###### Father and Mother #############

v1 = ws.range("F8:F64").value
v1_name = ws.range("D8:D64").value
print(v1_name)
v3 = len(v1)
v2 = []
print(v3)
for i in range(0, v3):
    if v1[i] is not None:
        v2.append(v1[i].split(','))

print(v2)
Father = []
Mother = []
v3 = sheet1.max_row

for i in range(0, len(v2)):
    for j in range(0, 2):
        if j == 0:
            Father.append(v2[i][j])
        if j == 1:
            Mother.append(v2[i][j])


count = 0
j = 1
for i in range(0, len(v1_name)):
    if v1_name[i] is not None:
        sheet1.cell(row=v3+j, column=16).value = Father[count]
        count += 1
        j += 1
    elif v1_name[i] is None:
        j += 1
        continue
j = 1
count1 = 0
for i in range(0, len(v1_name)):
    if v1_name[i] is not None:
        sheet1.cell(row=v3+j, column=17).value = Mother[count1]
        count1 += 1
        j += 1
    elif v1_name[i] is None:
        j += 1
        continue

v3_new = v3
################### END ######################





#####  FOR NAME #######
v1 = ws.range("D8:D64").value
print("Result:", v1)

v2 = len(v1)
v3 = v3_new

print(v3, v2)
count = 0
for i in range(v3, v3+v2):
    sheet1.cell(row=i+1, column=15).value = v1[count]
    count += 1

v3_new = v3
###### NAME END #######



#####  FOR STUDENT_ID #######
v1 = ws.range("B8:B64").value
print("Result:", v1)

v2 = len(v1)
v3 = v3_new


print(v3, v2)
count = 0
for i in range(v3, v3+v2):
    sheet1.cell(row=i+1, column=23).value = v1[count]
    count += 1
v3_new = v3
#####  END STUDENT_ID #######


#####  FOR Gender #######
v1 = ws.range("E8:E64").value
print("Result:", v1)

v2 = len(v1)
v3 = v3_new

print(v3, v2)
count = 0
for i in range(v3, v3+v2):
    sheet1.cell(row=i+1, column=13).value = v1[count]
    count += 1
v3_new = v3
#####  END Gender #######


#####  FOR SSC_YEAR #######
v1 = ws.range("O8:O64").value
print("Result:", v1)

v2 = len(v1)
v3 = v3_new

print(v3, v2)
count = 0
for i in range(v3, v3+v2):
    sheet1.cell(row=i+1, column=11).value = v1[count]
    count += 1
v3_new = v3
#####  END SSC_YEAR #######




#####  FOR SSC_BOARD #######
v1 = ws.range("J8:J64").value
print("Result:", v1)

v2 = len(v1)
v3 = v3_new

print(v3, v2)
count = 0
for i in range(v3, v3+v2):
    sheet1.cell(row=i+1, column=9).value = v1[count]
    count += 1
v3_new = v3
#####  END SSC_BOARD #######



#####  FOR HSC_YEAR #######
v1 = ws.range("P8:P64").value
print("Result:", v1)

v2 = len(v1)
v3 = v3_new

print(v3, v2)
count = 0
for i in range(v3, v3+v2):
    sheet1.cell(row=i+1, column=7).value = v1[count]
    count += 1
v3_new = v3
#####  END HSC_YEAR #######


#####  FOR HSC_BOARD #######
v1 = ws.range("K8:K64").value
print("Result:", v1)

v2 = len(v1)
v3 = v3_new

print(v3, v2)
count = 0
for i in range(v3, v3+v2):
    sheet1.cell(row=i+1, column=4).value = v1[count]
    count += 1
v3_new = v3
#####  END hSC_BOARD #######



#####  FOR SSC_GPA #######
v1 = ws.range("L8:L64").value
print("Result:", v1)

v2 = len(v1)
v3 = v3_new

print(v3, v2)
count = 0
for i in range(v3, v3+v2):
    sheet1.cell(row=i+1, column=18).value = v1[count]
    count += 1
v3_new = v3
#####  END SSC_GPA #######



#####  FOR HSC_GPA #######
v1 = ws.range("M8:M64").value
print("Result:", v1)

v2 = len(v1)
v3 = v3_new

print(v3, v2)
count = 0
for i in range(v3, v3+v2):
    sheet1.cell(row=i+1, column=19).value = v1[count]
    count += 1
v3_new = v3
#####  END HSC_GPA #######


#####  FOR RELIGION #######
v1 = ws.range("I8:I64").value
print("Result:", v1)

v2 = len(v1)
v3 = v3_new

print(v3, v2)
count = 0
for i in range(v3, v3+v2):
    sheet1.cell(row=i+1, column=46).value = v1[count]
    count += 1
v3_new = v3
#####  END RELIGION #######


#####  FOR SEMESTER ID  #######

v1 = ws.range("I8:I64").value
print("Result:", v1)

v2 = len(v1)
v3 = v3_new

print(v3, v2)
#count = 0
for i in range(v3, v3+v2):
    sheet1.cell(row=i+1, column=1).value = "11022007"
    #count += 1
v3_new = v3

#####  END SEMESTER ID  #######

XL.save('F:\Work\TABLE TEMPLATE\ADMISSION_STUDENTS.xlsx')

