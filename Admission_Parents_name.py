import xlwings as xw
import openpyxl as xl


ws = xw.Book("F:\Work\DATASET\Details Data Final Fall-2007.xls").sheets['ARC']
XL = xl.load_workbook("F:\Work\TABLE TEMPLATE\ADMISSION_STUDENTS.xlsx")
sheet1 = XL["Result 1"]

v1 = ws.range("G8:G46").value
v1_name = ws.range("D8:D46").value
print(v1_name)
v3 = len(v1)
v2 = []
print(v3)
for i in range(0, v3):
    if v1[i] is not None:
        v2.append(v1[i].split(','))

#print(v2)
Father = []
Mother = []


for i in range(0, len(v2)):
    for j in range(0, 2):
        if j == 0:
            Father.append(v2[i][j])
        if j == 1:
            Mother.append(v2[i][j])


count = 0
for i in range(0, len(v1_name)):
    if v1_name[i] is not None:
        sheet1.cell(row=i+3, column=16).value = Father[count]
        count += 1
    elif v1_name[i] is None:
        continue

count1 = 0
for i in range(0, len(v1_name)):
    if v1_name[i] is not None:
        sheet1.cell(row=i+3, column=17).value = Mother[count1]
        count1 += 1
    elif v1_name[i] is None:
        continue
XL.save('F:\Work\TABLE TEMPLATE\ADMISSION_STUDENTS.xlsx')