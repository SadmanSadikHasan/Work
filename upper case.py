import xlwings as xw
import openpyxl as xl

ws = xw.Book("F:\Work\TABLE TEMPLATE\ADMISSION_STUDENTS - Copy.xlsx").sheets['Result 1']
XL = xl.load_workbook("F:\Work\TABLE TEMPLATE\ADMISSION_STUDENTS.xlsx")
sheet1 = XL["Result 1"]

v1 = ws.range("O3:O299").value
v2 = ws.range("P3:P299").value
v3 = ws.range("Q3:Q299").value
v4 = ws.range("I3:I299").value
v5 = ws.range("D4:D299").value
v6 = ws.range("AW4:AW299").value

for i in range(0, len(v1)):
    if v1[i] is not None:
        v1[i] = v1[i].upper()
print("Result:", v1)

for i in range(0, len(v2)):
    if v2[i] is not None:
        v2[i] = v2[i].upper()
print("Result:", v2)

for i in range(0, len(v3)):
    if v3[i] is not None:
        v3[i] = v3[i].upper()
print("Result:", v3)

for i in range(0, len(v4)):
    if v4[i] is not None:
        v4[i] = v4[i].upper()
print("Result:", v4)

for i in range(0, len(v5)):
    if v5[i] is not None:
        v5[i] = v5[i].upper()
print("Result:", v5)

for i in range(0, len(v6)):
    if v6[i] is not None:
        v6[i] = v6[i].upper()
print("Result:", v6)

for i in range(0, len(v1)):
    sheet1.cell(row=i+3, column=15).value = v1[i]

for i in range(0, len(v2)):
    sheet1.cell(row=i+3, column=16).value = v2[i]

for i in range(0, len(v3)):
    sheet1.cell(row=i+3, column=17).value = v3[i]

for i in range(0, len(v4)):
    sheet1.cell(row=i+3, column=9).value = v4[i]

for i in range(0, len(v5)):
    sheet1.cell(row=i+3, column=4).value = v5[i]

for i in range(0, len(v6)):
    sheet1.cell(row=i+3, column=49).value = v6[i]

XL.save('F:\Work\TABLE TEMPLATE\ADMISSION_STUDENTS.xlsx')
